from __future__ import annotations

import html
import io
import json
import re
from datetime import datetime
from typing import Any

import pymupdf
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


PRIORITY_LABEL = {"high": "高", "medium": "中", "low": "低"}
KIND_LABEL = {"modified": "修改", "added": "新增", "deleted": "刪除"}
VERDICT_LABEL = {
    "confirmed": "確認有變更",
    "false_positive": "誤報",
    "insufficient_evidence": "證據不足",
}
DISPOSITION_LABEL = {
    "no_action": "無需處置",
    "follow_up": "需要追蹤",
    "action_complete": "已完成處置",
}


def safe(value: Any) -> str:
    return html.escape(str(value or ""))


def page_label(value: Any) -> str:
    return f"p.{value}" if value else "—"


def report_title(session: dict[str, Any]) -> str:
    return f"{session['old_name']} 與 {session['new_name']} 版本差異報告"


def build_html_report(session: dict[str, Any], printable: bool = False) -> bytes:
    state_label = "正式審查報告" if session["report_state"] == "official" else "修訂草稿"
    rows = []
    for diff in session["differences"]:
        verdict = VERDICT_LABEL.get(diff.get("verdict"), "未覆核")
        disposition = DISPOSITION_LABEL.get(diff.get("disposition"), "—")
        explanation = diff.get("llm_explanation") or diff.get("system_explanation") or ""
        rows.append(
            f"""
            <article class="diff {safe(diff['priority'])}">
              <header><span>{safe(diff['id'])}</span><strong>{safe(diff['title'])}</strong>
                <em>{PRIORITY_LABEL.get(diff['priority'], diff['priority'])}優先</em></header>
              <div class="meta">{KIND_LABEL.get(diff['kind'], diff['kind'])} · 舊 {page_label(diff.get('old_page'))} → 新 {page_label(diff.get('new_page'))} · 信心度 {safe(diff['confidence'])}</div>
              <div class="columns"><section><h3>舊版</h3><p>{safe(diff.get('old_text') or '（無對應內容）')}</p></section>
              <section><h3>新版</h3><p>{safe(diff.get('new_text') or '（無對應內容）')}</p></section></div>
              <p><b>系統整理：</b>{safe(explanation)}</p>
              <p><b>人工判定：</b>{safe(verdict)}　<b>處置：</b>{safe(disposition)}　<b>覆核人：</b>{safe(diff.get('reviewer') or '—')}</p>
              <p><b>覆核備註：</b>{safe(diff.get('note') or '—')}</p>
            </article>
            """
        )

    coverage_rows = "".join(
        f"<li>{safe(issue['side'].upper())} {page_label(issue['page_number'])}：{safe(issue['message'])}"
        f"（{'已確認' if issue['status'] == 'acknowledged' else '待確認'}"
        f"{('，' + safe(issue.get('resolved_by'))) if issue.get('resolved_by') else ''}）</li>"
        for issue in session.get("coverage", [])
    ) or "<li>沒有頁面覆蓋問題。</li>"

    auto_print = "<script>window.addEventListener('load',()=>setTimeout(()=>window.print(),250));</script>" if printable else ""
    document = f"""<!doctype html>
<html lang="zh-Hant"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>{safe(report_title(session))}</title>
<style>
@page{{size:A4;margin:14mm}}*{{box-sizing:border-box}}body{{margin:0;background:#eef3f7;color:#172536;font:13px/1.65 Inter,"Noto Sans TC",system-ui,sans-serif}}
.report{{max-width:980px;margin:24px auto;background:#fff;padding:36px;box-shadow:0 8px 30px #16304718}}
h1{{font-size:24px;margin:0 0 4px}}.state{{display:inline-block;padding:5px 10px;border-radius:999px;background:{'#e8f6ef' if session['report_state']=='official' else '#fff3cf'};font-weight:750}}
.summary{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin:24px 0}}.metric{{border:1px solid #dbe4eb;border-radius:8px;padding:12px}}.metric b{{display:block;font-size:22px;color:#0a3e63}}
.identity{{border-collapse:collapse;width:100%;margin:18px 0}}.identity td{{border:1px solid #dbe4eb;padding:7px 9px}}
.diff{{border:1px solid #dbe4eb;border-left:5px solid #8291a0;border-radius:8px;margin:14px 0;padding:14px;break-inside:avoid}}.diff.high{{border-left-color:#c63a32}}.diff.medium{{border-left-color:#d99a21}}.diff.low{{border-left-color:#2b8a62}}
.diff header{{display:flex;gap:10px;align-items:center}}.diff header strong{{flex:1}}.diff header em{{font-style:normal;font-weight:750}}.meta{{color:#657486;margin:5px 0 10px}}
.columns{{display:grid;grid-template-columns:1fr 1fr;gap:10px}}.columns section{{background:#f6f8fa;border-radius:6px;padding:9px}}.columns h3{{font-size:11px;text-transform:uppercase;color:#647384;margin:0 0 4px}}.columns p{{white-space:pre-wrap;margin:0}}
footer{{border-top:1px solid #dbe4eb;margin-top:24px;padding-top:12px;color:#657486}}@media print{{body{{background:#fff}}.report{{box-shadow:none;margin:0;padding:0}}}}
</style>{auto_print}</head><body><main class="report">
<span class="state">{state_label}</span><h1>{safe(report_title(session))}</h1>
<p>報告編號：PLM-{safe(session['id'].upper())}　建立時間：{safe(session['created_at'])}</p>
<table class="identity"><tr><td><b>舊版</b></td><td>{safe(session['old_name'])}</td><td>{safe(session.get('old_version') or '未填')}</td><td>SHA-256 {safe(session['old_sha256'][:16])}…</td></tr>
<tr><td><b>新版</b></td><td>{safe(session['new_name'])}</td><td>{safe(session.get('new_version') or '未填')}</td><td>SHA-256 {safe(session['new_sha256'][:16])}…</td></tr></table>
<div class="summary"><div class="metric">總差異<b>{session['total_differences']}</b></div><div class="metric">高優先<b>{session['high_count']}</b></div><div class="metric">已覆核<b>{session['reviewed_count']}</b></div><div class="metric">待處理關卡<b>{session['must_review_open']}</b></div></div>
<h2>比對覆蓋狀態</h2><ul>{coverage_rows}</ul><h2>差異明細</h2>{''.join(rows)}
<footer>本報告由 Plimsoll 產生。系統優先級與 LLM 解讀只供覆核輔助，最終判定仍由具權責人員負責。</footer>
</main></body></html>"""
    return document.encode("utf-8")


def build_xlsx_report(session: dict[str, Any]) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "摘要"
    summary.append(["Plimsoll 船舶技術文件版本差異報告"])
    summary["A1"].font = Font(size=18, bold=True, color="173B56")
    summary.append(["報告狀態", "正式審查報告" if session["report_state"] == "official" else "修訂草稿"])
    summary.append(["工作階段", session["id"]])
    summary.append(["舊版文件", session["old_name"], session.get("old_version")])
    summary.append(["新版文件", session["new_name"], session.get("new_version")])
    summary.append(["總差異", session["total_differences"]])
    summary.append(["高／中／低", session["high_count"], session["medium_count"], session["low_count"]])
    summary.append(["已覆核", session["reviewed_count"]])
    summary.append(["待處理關卡", session["must_review_open"]])

    sheet = workbook.create_sheet("差異清單")
    headers = [
        "差異 ID", "候選優先級", "種類", "標題", "舊頁", "新頁", "舊版原文",
        "新版原文", "觸發規則", "信心度", "人工判定", "處置", "最終優先級",
        "覆核人", "覆核備註", "覆核時間",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="173B56")
    for cell in sheet[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for diff in session["differences"]:
        sheet.append(
            [
                diff["id"], PRIORITY_LABEL.get(diff["priority"], diff["priority"]),
                KIND_LABEL.get(diff["kind"], diff["kind"]), diff["title"],
                diff.get("old_page"), diff.get("new_page"), diff.get("old_text"),
                diff.get("new_text"), "、".join(diff.get("triggers", [])),
                diff["confidence"], VERDICT_LABEL.get(diff.get("verdict"), "未覆核"),
                DISPOSITION_LABEL.get(diff.get("disposition"), ""),
                PRIORITY_LABEL.get(diff.get("final_priority"), diff.get("final_priority") or ""),
                diff.get("reviewer"), diff.get("note"), diff.get("reviewed_at"),
            ]
        )
    widths = [12, 12, 10, 34, 9, 9, 54, 54, 24, 10, 14, 14, 12, 14, 34, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    coverage = workbook.create_sheet("頁面覆蓋問題")
    coverage.append(["文件側", "頁碼", "問題類型", "說明", "狀態", "確認人", "確認備註", "確認時間"])
    for issue in session.get("coverage", []):
        coverage.append(
            [
                issue["side"], issue["page_number"], issue["issue_type"], issue["message"],
                issue["status"], issue.get("resolved_by"), issue.get("resolution_note"),
                issue.get("resolved_at"),
            ]
        )
    for cell in coverage[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
    coverage.column_dimensions["A"].width = 12
    coverage.column_dimensions["B"].width = 10
    coverage.column_dimensions["C"].width = 22
    coverage.column_dimensions["D"].width = 70
    coverage.column_dimensions["E"].width = 12
    coverage.column_dimensions["F"].width = 16
    coverage.column_dimensions["G"].width = 40
    coverage.column_dimensions["H"].width = 22

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _wrap_text(text: str, width: int = 46) -> list[str]:
    text = re.sub(r"\s+", " ", text or "").strip()
    if not text:
        return ["—"]
    return [text[index : index + width] for index in range(0, len(text), width)]


def build_pdf_report(session: dict[str, Any]) -> bytes:
    document = pymupdf.open()
    page = None
    y = 0.0

    def new_page() -> None:
        nonlocal page, y
        page = document.new_page(width=595, height=842)
        y = 48
        page.insert_text((48, 28), "Plimsoll 船舶技術文件版本差異報告", fontname="china-t", fontsize=8, color=(0.3, 0.4, 0.5))

    def line(text: str, size: float = 10, bold: bool = False, gap: float = 4, color: tuple[float, float, float] = (0.08, 0.14, 0.2)) -> None:
        nonlocal y
        if page is None:
            new_page()
        chunks = _wrap_text(text, 34 if size >= 16 else 52)
        for chunk in chunks:
            if y > 798:
                new_page()
            page.insert_text((48, y), chunk, fontname="china-t", fontsize=size, color=color)
            y += size * 1.45
        y += gap

    new_page()
    state = "正式審查報告" if session["report_state"] == "official" else "修訂草稿"
    line(state, 11, color=(0.72, 0.25, 0.08) if session["report_state"] != "official" else (0.08, 0.48, 0.3))
    line(report_title(session), 18)
    line(f"報告編號：PLM-{session['id'].upper()}　建立時間：{session['created_at']}", 9)
    line(f"舊版：{session['old_name']}　{session.get('old_version') or '未填'}", 10)
    line(f"新版：{session['new_name']}　{session.get('new_version') or '未填'}", 10)
    line(
        f"差異 {session['total_differences']} 筆｜高 {session['high_count']}｜中 {session['medium_count']}｜低 {session['low_count']}｜已覆核 {session['reviewed_count']}",
        11,
    )
    if session.get("coverage"):
        line("頁面覆蓋問題", 14)
        for issue in session["coverage"]:
            status = "已人工確認" if issue["status"] == "acknowledged" else "待人工確認"
            reviewer = f"｜{issue.get('resolved_by')}" if issue.get("resolved_by") else ""
            line(
                f"• {issue['side'].upper()} p.{issue['page_number']}：{issue['message']}｜{status}{reviewer}",
                9,
            )

    line("差異明細", 14)
    for diff in session["differences"]:
        line(
            f"{diff['id']}｜{PRIORITY_LABEL.get(diff['priority'], diff['priority'])}優先｜{diff['title']}",
            11,
            color=(0.68, 0.14, 0.12) if diff["priority"] == "high" else (0.08, 0.18, 0.27),
        )
        line(f"舊 {page_label(diff.get('old_page'))}：{(diff.get('old_text') or '（無對應內容）')[:260]}", 9, gap=1)
        line(f"新 {page_label(diff.get('new_page'))}：{(diff.get('new_text') or '（無對應內容）')[:260]}", 9, gap=1)
        line(
            f"人工判定：{VERDICT_LABEL.get(diff.get('verdict'), '未覆核')}｜處置：{DISPOSITION_LABEL.get(diff.get('disposition'), '—')}｜覆核人：{diff.get('reviewer') or '—'}",
            9,
            gap=8,
        )

    line("系統優先級與 LLM 解讀只供覆核輔助，最終判定仍由具權責人員負責。", 8, color=(0.4, 0.46, 0.52))
    return document.tobytes(garbage=4, deflate=True)


def build_jsonl_export(session: dict[str, Any]) -> bytes:
    records = []
    for diff in session["differences"]:
        record = {
            "session_id": session["id"],
            "difference_id": diff["id"],
            "kind": diff["kind"],
            "candidate_priority": diff["priority"],
            "confidence": diff["confidence"],
            "old_page": diff.get("old_page"),
            "new_page": diff.get("new_page"),
            "old_text": diff.get("old_text"),
            "new_text": diff.get("new_text"),
            "triggers": diff.get("triggers", []),
            "review": {
                "verdict": diff.get("verdict"),
                "disposition": diff.get("disposition"),
                "final_priority": diff.get("final_priority"),
                "reviewer": diff.get("reviewer"),
                "note": diff.get("note"),
                "reviewed_at": diff.get("reviewed_at"),
            },
        }
        records.append(json.dumps(record, ensure_ascii=False))
    return ("\n".join(records) + "\n").encode("utf-8")
